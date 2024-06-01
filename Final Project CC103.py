from cProfile import label
from ctypes import resize
from email.mime import image
from logging import root
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#EXCEL=================================================================================================================================================================================================================================

excel_con = Workbook()

excel_con = load_workbook('FINAL.xlsx')

excel_activate = excel_con.active


#GUI==================================================================================================================================================================================================================================

root= tk.Tk()
root.geometry('860x490')
root.title('WELCOME TO KENJI RESORT')
root.configure(bg='lightblue')


#WELCOME LABEL=========================================================================================================================================================================================================================

welcome= Label(root, text= "W E L C O M E  T O  K E N J I   R E S O R T",font =("Arial", 15, "bold"),bg ='lightblue', bd= 5 )
welcome.pack()

#PICTURE NG WELCOME LABEL===============================================================================================================================================================================================================

img= ImageTk.PhotoImage(Image.open('welcome.jpg'))
label= Label(root,image=img,bd=4)
label.pack()

#NEWROOT================================================================================================================================================================================================================================
def reserve():
    NewRoot = tk.Toplevel()
    NewRoot.geometry('1900x900')
    NewRoot.title('COTTAGE RESERVATION')
    NewRoot.configure(bg="grey")
    NewRoot.resizable(False,False)
    root.withdraw()

    img= ImageTk.PhotoImage(Image.open('image.jpg'))
    label= Label(NewRoot,image=img,bd=4)
    label.pack()

#FRAMES====================================================================================================================================================================================================================================

    coverframe = Frame(NewRoot, width = 1000, height= 570, bg ='skyblue', borderwidth=3 , relief="solid")
    oneframe = Frame(NewRoot, width = 490, height= 550, bg ='lightgray', borderwidth=3 , relief="solid")
    twoframe = Frame(NewRoot, width = 460, height= 480, bg ='lightgray', borderwidth=3 , relief="solid")
        
#BUTTON FUNCTIONS============================================================================================================================================================================================================================

    def reset():
        refe_input.delete(0,END)
        name_input.delete(0,END)
        mail_input.delete(0,END)
        gender_var.set('Male')
        desi_input.delete(0,END)
        cont_input.delete(0,END)
        sala_input.delete(0,END)
        addr_input.delete("0.1","end")

    def delete():
        if len(name_input.get()) == 0:
            messagebox.showerror("DELETE","INPUT YOUR NAME FIRST")
            
        else:
            found = False
            for each_cell in range(2, (excel_activate.max_row)+1):
                if (name_input.get() ==  excel_activate['B'+str(each_cell)].value):
                    #Found is a boolean variable that will trigger the stopping point of the loop 
                    found = True
                    cell_address = each_cell
                    break
            if found:
                excel_activate.delete_rows(cell_address)
                messagebox.showinfo("INFO","DATA DELETED")
            else:
                messagebox.showerror("DELETE","THESE NAME IS NOT EXISTED IN OUR DATA")
                    
                
            excel_con.save('FINAL.xlsx')
        
    def save():
        emp_refer = refe_input.get()
        emp_nam = name_input.get()
        emp_ema = mail_input.get()
        emp_gen = gender_var.get()
        emp_des = desi_input.get()
        emp_con = cont_input.get()
        emp_sal = sala_input.get()
        emp_add = addr_input.get('0.1','end-1c')

        found = False
        for each_cell in range(1, excel_activate.max_row + 1):
            if emp_nam == excel_activate['B' + str(each_cell)].value or emp_con == excel_activate['f' + str(each_cell)].value:
                found = True
                
                break
            
        if found == True:
            messagebox.showerror("DATA","This data is already existed")

        else:
            lastrow = str(excel_activate.max_row+1)
            excel_activate['A'+lastrow] = emp_refer
            excel_activate['B'+lastrow] = emp_nam
            excel_activate['C'+lastrow] = emp_ema
            excel_activate['D'+lastrow] = emp_gen
            excel_activate['E'+lastrow] = emp_des
            excel_activate['F'+lastrow] = emp_con
            excel_activate['G'+lastrow] = emp_sal
            excel_activate['H'+lastrow] = emp_add

            

            messagebox.showinfo("DATABASE RECORDS","DATA SAVED SUCCESSFULLY")
            
            excel_con.save('FINAL.xlsx')
            

    def search():
        if len(name_input.get()) == 0:
            messagebox.showerror("SEARCH","INPUT YOUR NAME FIRST")
        else:
            Found = False
            for each_cell in range(2, (excel_activate.max_row)+1):
                if (name_input.get() ==  excel_activate['B'+str(each_cell)].value):
                        #Found is a boolean variable that will trigger the stopping point of the loop 
                    Found = True
                    cell_address = str(each_cell)
                    break
                else:
                    Found=False
            if(Found == True):
                messagebox.showinfo("FOUND","DATA EXIST IN " + cell_address )
            else:
                messagebox.showerror("NOT FOUND"," DATA NOT EXIST ")

    def exit():
        messagebox.showinfo('RESERVATION SUCCESFUL','YOU HAVE BEEN RESERVE YOUR TIME ON RESORT:''\n THANKYOU and HAVE A NICE DAY!')
        NewRoot.withdraw()
        excel_con.save('FINAL.xlsx')
        root.deiconify()
    
        

    def editframe():
        namefound = False
        for each_cell in range(2, excel_activate.max_row+1):
                if name_input.get() == excel_activate['B' + str(each_cell)].value:
                    namefound = True
                    break
        if namefound == False:
            messagebox.showerror('ERROR','NAME DOES NOT EXIST')

        found = False
        if len(name_input.get()) == 0:
            messagebox.showerror('ERROR','INPUT YOUR NAME FIRST TO LOCATE THE DATA')
            
        else:
            for each_cell in range(2, excel_activate.max_row+1):
                if name_input.get() == excel_activate['B' + str(each_cell)].value:
                    found = True
                    current = each_cell
                    break
            if found == True:
                edit_frame = tk.Toplevel()
                edit_frame.geometry('600x400')
                edit_frame.title('Data Update')
                edit_frame.configure(bg='grey')

                img= ImageTk.PhotoImage(Image.open('pool.jpg'))
                label= Label(edit_frame,image=img,bd=4)
                label.pack()


                editframe = Frame(edit_frame, width = 380, height= 300, bg ='lightgrey', borderwidth=3 , relief="solid")
                editframe.place(x=80,y=50)

                emplExcel = StringVar()
                nameExcel = StringVar()
                mailExcel = StringVar()
                gendExcel = StringVar()
                desiExcel = StringVar()
                contExcel = StringVar()
                salaExcel = StringVar()
                addrExcel = StringVar()


                edit_label = Label(editframe, text= 'Edit Information',bg='beige',font=('arial',20,'bold'),relief="solid")
                edit_label.pack()

                    # Student No. =====================================================================
                empl_label = Label(editframe, text= 'Reservation No.:',font=('arial',10,'bold'),bg='lightgrey')
                empl_ent = Entry(editframe, width= 20, textvariable=emplExcel)
                empl_choice = IntVar()
                emplChk = Checkbutton(editframe, text="same as before", variable=empl_choice, command=lambda:get_existing_empl(),bg='lightgrey')

                    # Name =====================================================================
                name_label = Label(editframe, text= 'Name :',font=('arial',10,'bold'),bg='lightgrey')
                name_ent = Entry(editframe, width= 20, textvariable=nameExcel)
                name_choice = IntVar()
                nameChk = Checkbutton(editframe, text="same as before", variable=name_choice, command=lambda:get_existing_name(),bg='lightgrey')
                    

                    # Gmail =====================================================================
                mail_label = Label(editframe, text= 'Email :',font=('arial',10,'bold'),bg='lightgrey')
                mail_ent = Entry(editframe, width= 20, textvariable=mailExcel)
                mail_choice = IntVar()
                mailChk = Checkbutton(editframe, text="same as before", variable=mail_choice, command=lambda:get_existing_mail(),bg='lightgrey')

                    # Gender =====================================================================
                gend_label = Label(editframe, text= 'Overnight:',font=('arial',10,'bold'),bg='lightgrey')
                gend_ent = Entry(editframe, width= 20, textvariable=gendExcel)
                gend_choice = IntVar()
                gendChk = Checkbutton(editframe, text="same as before", variable=gend_choice, command=lambda:get_existing_gend(),bg='lightgrey')

                    # Year/Section =====================================================================
                desi_label = Label(editframe, text= 'Cottage Size:',font=('arial',10,'bold'),bg='lightgrey')
                desi_ent = Entry(editframe, width= 20, textvariable=desiExcel)
                desi_choice = IntVar()
                desiChk = Checkbutton(editframe, text="same as before", variable=desi_choice, command=lambda:get_existing_desi(),bg='lightgrey')

                    # Contact No. =====================================================================
                cont_label = Label(editframe, text= 'Contact no :',font=('arial',10,'bold'),bg='lightgrey')
                cont_ent = Entry(editframe, width= 20, textvariable=contExcel)
                cont_choice = IntVar()
                contChk = Checkbutton(editframe, text="same as before", variable=cont_choice, command=lambda:get_existing_cont(),bg='lightgrey')

                    # Birthdate =====================================================================
                sala_label = Label(editframe, text= 'Time Of Visit:',font=('arial',10,'bold'),bg='lightgrey')
                sala_ent = Entry(editframe, width= 20, textvariable=salaExcel)
                sala_choice = IntVar()
                salaChk = Checkbutton(editframe, text="same as before", variable=sala_choice, command=lambda:get_existing_sala(),bg='lightgrey')

                    # Address =====================================================================
                addr_label = Label(editframe, text= 'Message :',font=('arial',10,'bold'),bg='lightgrey')
                addr_ent = Entry(editframe, width= 20, textvariable=addrExcel)
                addr_choice = IntVar()
                addrChk = Checkbutton(editframe, text="same as before", variable=addr_choice, command=lambda:get_existing_addr(),bg='lightgrey')

                    
                updatebtn = Button(editframe, text ="Update",bg='skyblue',font=('arial',15,'bold'),command=lambda:update())

                def get_existing_empl():
                    if empl_choice.get() == 1:
                        emplOld = empl
                        emplExcel.set(emplOld)
                    elif empl_choice.get() ==0:
                        emplExcel.set("")

                def get_existing_name():
                    if name_choice.get() == 1:
                        nameOld = name
                        nameExcel.set(nameOld)
                    elif name_choice.get() ==0:
                        nameExcel.set("")

                def get_existing_mail():
                    if mail_choice.get() == 1:
                        mailOld = mail
                        mailExcel.set(mailOld)
                    elif mail_choice.get() ==0:
                        mailExcel.set("")

                def get_existing_gend():
                    if gend_choice.get() == 1:
                        gendOld = gend
                        gendExcel.set(gendOld)
                    elif gend_choice.get() ==0:
                        gendExcel.set("")

                def get_existing_desi():
                    if desi_choice.get() == 1:
                        desiOld = desi
                        desiExcel.set(desiOld)
                    elif desi_choice.get() ==0:
                        desiExcel.set("")

                def get_existing_cont():
                    if cont_choice.get() == 1:
                        contOld = cont
                        contExcel.set(contOld)
                    elif name_choice.get() ==0:
                        nameExcel.set("")

                def get_existing_sala():
                    if sala_choice.get() == 1:
                        salaOld = sala
                        salaExcel.set(salaOld)
                    elif sala_choice.get() ==0:
                        salaExcel.set("")

                def get_existing_addr():
                    if addr_choice.get() == 1:
                        addrOld = addr
                        addrExcel.set(addrOld)
                    elif addr_choice.get() ==0:
                        addrExcel.set("")
                    
                def update():
                    excel_activate['A'+ str(each_cell)].value = empl_ent.get()
                    excel_activate['B'+ str(each_cell)].value = name_ent.get()
                    excel_activate['C'+ str(each_cell)].value = mail_ent.get()
                    excel_activate['D'+ str(each_cell)].value = gend_ent.get()
                    excel_activate['E'+ str(each_cell)].value = desi_ent.get()
                    excel_activate['F'+ str(each_cell)].value = cont_ent.get()
                    excel_activate['G'+ str(each_cell)].value = sala_ent.get()
                    excel_activate['H'+ str(each_cell)].value = addr_ent.get()

                    excel_con.save('Excel_DATA.xlsx')
                    messagebox.showinfo("UPDATED","DATA HAS BEEN UPDATED")
                    editframe.withdraw()

                empl =excel_activate['A'+ str(current)].value
                name =excel_activate['B'+ str(current)].value
                mail =excel_activate['C'+ str(current)].value
                gend=excel_activate['D'+ str(current)].value
                desi =excel_activate['E'+ str(current)].value
                cont =excel_activate['F'+ str(current)].value
                sala =excel_activate['G'+ str(current)].value
                addr =excel_activate['H'+ str(current)].value

                edit_label.grid(row = 0 , column= 1)
                    #label
                empl_label.grid(row = 1, column = 0)
                name_label.grid(row = 2, column = 0)
                mail_label.grid(row = 3, column = 0)
                gend_label.grid(row = 4, column = 0)
                desi_label.grid(row = 5, column = 0)
                cont_label.grid(row = 6, column = 0)
                sala_label.grid(row = 7, column = 0)
                addr_label.grid(row = 8, column = 0)

                    #Entry
                empl_ent.grid(row = 1, column = 1)
                name_ent.grid(row = 2, column = 1)
                mail_ent.grid(row = 3, column = 1)
                gend_ent.grid(row = 4, column = 1)
                desi_ent.grid(row = 5, column = 1)
                cont_ent.grid(row = 6, column = 1)
                sala_ent.grid(row = 7, column = 1)
                addr_ent.grid(row = 8, column = 1)

                    #Checkbutton
                emplChk.grid(row = 1, column =2)
                nameChk.grid(row = 2, column =2)
                mailChk.grid(row = 3, column =2)
                gendChk.grid(row = 4, column =2)
                desiChk.grid(row = 5, column =2)
                contChk.grid(row = 6, column =2)
                salaChk.grid(row = 7, column =2)
                addrChk.grid(row = 8, column =2)
                    




                updatebtn.grid(row = 9, column= 1)

                editframe.mainloop()

                    
    def view():
        view_frame = Toplevel()
        view_frame.geometry('1190x350')
        view_frame.title('LIST OF RESERVATIONS')
        view_frame.configure(bg='grey')
        view_frame.resizable(False,False)

        img= ImageTk.PhotoImage(Image.open('swimm.jpg'))
        label= Label(view_frame,image=img,bd=4)
        label.pack()

        vframe = Frame(view_frame, width = 380, height= 500, bg ='lightgrey', borderwidth=3 , relief="solid")
        vframe.place(x=45,y=50)

        tv1 = ttk.Treeview(vframe)
        treescrolly = Scrollbar(vframe,orient="vertical", command=tv1.yview)
        treescrollx = Scrollbar(vframe,orient="horizontal", command=tv1.xview)
        tv1.configure(xscrollcommand = treescrollx.set, yscrollcommand = treescrolly.set)
        treescrollx.pack(side = "bottom", fill="x")
        treescrolly.pack(side = "right", fill="y")

        tv1['columns'] = ("Reservation No.", "Fullname", "Email", "Overnight", "Cottage Size", "Contact No.", "Time of Visit:", "Message:")
        tv1.column("#0", width= 120, minwidth =25)
        tv1.column("Reservation No.", anchor= W, width = 120)
        tv1.column("Fullname", anchor= W, width = 120)
        tv1.column("Email", anchor= W, width = 120)
        tv1.column("Overnight", anchor= W, width = 120)
        tv1.column("Cottage Size", anchor= W, width = 120)
        tv1.column("Contact No.", anchor= W, width = 120)
        tv1.column("Time of Visit:", anchor= W, width = 120)
        tv1.column("Message:", anchor= W, width = 120)
                
        tv1.heading("#0", text="CUSTOMER NO.", anchor=W)
        tv1.heading("Reservation No.", text="Reservation No.", anchor= W)
        tv1.heading("Fullname", text="Fullname", anchor= W)
        tv1.heading("Email", text="Email", anchor= W)
        tv1.heading("Overnight", text="Overnight", anchor= W)
        tv1.heading("Cottage Size", text="Cottage Size", anchor= W)
        tv1.heading("Contact No.", text="Contact No.", anchor= W)
        tv1.heading("Time of Visit:", text="Time of Visit:", anchor= W)
        tv1.heading("Message:", text="Message:", anchor= W)


        for each_cell in range(2,(excel_activate.max_row)+1):
            tv1.insert(parent='', index="end", text=str(each_cell), values=(excel_activate["A"+str(each_cell)].value,excel_activate["B"+str(each_cell)].value,excel_activate["C"+str(each_cell)].value,excel_activate["D"+str(each_cell)].value,excel_activate["E"+str(each_cell)].value,excel_activate["F"+str(each_cell)].value,excel_activate["G"+str(each_cell)].value,excel_activate["H"+str(each_cell)].value))
        tv1.pack()
        view_frame.mainloop()




    # 2nd Interface -----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    welcome_frame= Frame(NewRoot, width = 1000, height= 570, bg ='skyblue', borderwidth=3 , relief="solid")
    welcome_frame.place(x = 600, y = 20)
    welcome = Label(welcome_frame, text= "K E N J I   RESORT   RESERVATION", font =("Arial", 30, "bold"),bg="white",fg= "black",relief="solid")


    # Loob ng frame 1 ----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    emp_refe= Label(oneframe, text= "RESERVATION NO.",font =("Arial", 12, "bold"),bg ='lightgray', bd= 5 )
    emp_name= Label(oneframe, text= "FULL NAME:",font =("Arial", 15, "bold"), bg ='lightgray', bd= 5)
    emp_mail= Label(oneframe, text= "EMAIL:",font =("Arial", 15, "bold"), bg ='lightgray', bd= 5)
    emp_gend= Label(oneframe, text= "OVERNIGHT:",font =("Arial", 15, "bold"), bg ='lightgray', bd= 5)
    emp_desi= Label(oneframe, text= "COTTAGE SIZE:",font =("Arial", 14, "bold"), bg ='lightgray', bd= 5)
    emp_cont= Label(oneframe, text= "CONTACT NO.",font =("Arial", 15, "bold"), bg ='lightgray', bd= 5)
    emp_sala= Label(oneframe, text= "TIME OF VISIT:",font =("Arial", 14, "bold"), bg ='lightgray', bd= 5)
    emp_addr= Label(oneframe, text= 'MESSAGE:',font =("Arial", 15, "bold"), bg ='lightgray', bd= 5)

    refe_input = Entry(oneframe, width = 40,font =("arial", 10, "bold"), bd= 5)
    name_input = Entry(oneframe, width = 40,font =("arial", 10, "bold"), bd= 5)
    mail_input = Entry(oneframe, width = 40,font =("arial", 10, "bold"), bd= 5)

    gender_var = StringVar()
    gender_var.set('None')
    male = Radiobutton(oneframe, text='YES',variable=gender_var, value = 'YES',font =("Arial", 15, "bold"), bg ='lightgray')
    female = Radiobutton(oneframe, text='NO',variable=gender_var, value = 'NO',font =("Arial", 15, "bold"), bg ='lightgray')
        
    desi_input = Entry(oneframe, width = 40,font =("arial", 10, "bold"), bd= 5)
    cont_input = Entry(oneframe, width = 40,font =("arial", 10, "bold"), bd= 5)
    sala_input = Entry(oneframe, width = 40,font =("arial", 10, "bold"), bd= 5)
    addr_input = Text(oneframe, width = 40, height= 5,font =("arial", 10, "bold"), bd= 5)

        #place for frame one =========================================================================================================================================================================================
    emp_refe.place(x= 3, y = 20)
    refe_input.place(x= 160, y = 15, height= 35)

    emp_name.place(x= 3, y = 70)
    name_input.place(x= 160, y = 65, height= 35)

    emp_mail.place(x= 3, y = 120)
    mail_input.place(x= 160, y = 115, height= 35)
    emp_gend.place(x= 3, y = 170)
    male.place(x= 160, y = 170, height= 35)
    female.place(x= 330, y = 170, height= 35)

    emp_desi.place(x= 3, y = 220)
    desi_input.place(x= 160, y = 215, height= 35)

    emp_cont.place(x= 3, y = 270)
    cont_input.place(x= 160, y = 265, height= 35)

    emp_sala.place(x= 3, y = 320)
    sala_input.place(x= 160, y = 320, height= 35)

    emp_addr.place(x= 3, y = 390)
    addr_input.place(x= 160, y = 385, height= 120)


        #loob ng frame 2 =====================================================================================================================================================================================================
    emp_details = Button(twoframe, text= "DETAILS",font =("Arial", 13, "bold"),bd=9,bg= "skyblue",command=lambda:view())
    btn11 = Button(twoframe, text = "ADD", font =("Arial", 13, "bold"),bg= "skyblue",fg = "black", command=lambda:save(), bd= 9)
    btn12 = Button(twoframe, text = "DELETE", font =("Arial", 13, "bold"),bg= "skyblue",fg = "black", command=lambda:delete(), bd= 9)
    btn13 = Button(twoframe, text = "SEARCH", font =("Arial", 13, "bold"),bg= "skyblue",fg = "black", command=lambda:search(), bd= 9)
    btn14 = Button(twoframe, text = "EDIT DATA", font =("Arial", 13, "bold"),bg= "skyblue",fg = "black",command=lambda:editframe(), bd= 9)
    btn15 = Button(twoframe, text = "EXIT", font =("Arial", 13, "bold"),bg= "skyblue",fg = "black",command=lambda:exit() , bd= 9)
    btn16 = Button(twoframe, text = "VIEW DATA", font =("Arial", 13, "bold"),bg= "skyblue",fg = "black",command=lambda:view() , bd= 9)
    btn17 = Button(twoframe, text = "RESET", font =("Arial", 13, "bold"),bg= "beige",fg = "Black",command=lambda:reset(), bd= 9 )

        #place for frame 2 ======================================================================================================================================================================================================

    emp_details.place(x= 3 ,y = 3, width= 449)
    btn11.place(x = 3, y = 60, height= 50,width= 449)
    btn12.place(x = 3, y = 120, height= 50,width= 449)
    btn13.place(x = 3, y = 180, height= 50,width= 449)
    btn14.place(x = 3, y = 240, height= 50,width= 449)
    btn15.place(x = 3, y = 300, height= 50,width= 449)
    btn16.place(x = 3, y = 360, height= 50,width= 449)
    btn17.place(x = 3, y = 420, height= 50,width= 449) 
   


#PLACE NG FRAME SA 2ND INTERFACE=========================================================================================================================================================================================================
    coverframe.place(x = 440, y = 100)
    oneframe.place(x = 460,y = 110)
    twoframe.place(x = 960, y = 140)

        
    welcome.pack()
        

    NewRoot.mainloop()   

#FOR ROOT INTERFACE==========================================================================================================================================================================================================================


def view2():
    view_frame = Toplevel()
    view_frame.geometry('1190x350')
    view_frame.title('LIST OF RESERVATIONS')
    view_frame.configure(bg='grey')
    view_frame.resizable(False,False)

    img= ImageTk.PhotoImage(Image.open('swimm.jpg'))
    label= Label(view_frame,image=img,bd=4)
    label.pack()

    vframe = Frame(view_frame, width = 380, height= 500, bg ='lightgrey', borderwidth=3 , relief="solid")
    vframe.place(x=45,y=50)

    tv1 = ttk.Treeview(vframe)
    treescrolly = Scrollbar(vframe,orient="vertical", command=tv1.yview)
    treescrollx = Scrollbar(vframe,orient="horizontal", command=tv1.xview)
    tv1.configure(xscrollcommand = treescrollx.set, yscrollcommand = treescrolly.set)
    treescrollx.pack(side = "bottom", fill="x")
    treescrolly.pack(side = "right", fill="y")

    tv1['columns'] = ("Reservation No.", "Fullname", "Email", "Overnight", "Cottage Size", "Contact No.", "Time of Visit:", "Message:")
    tv1.column("#0", width= 120, minwidth =25)
    tv1.column("Reservation No.", anchor= W, width = 120)
    tv1.column("Fullname", anchor= W, width = 120)
    tv1.column("Email", anchor= W, width = 120)
    tv1.column("Overnight", anchor= W, width = 120)
    tv1.column("Cottage Size", anchor= W, width = 120)
    tv1.column("Contact No.", anchor= W, width = 120)
    tv1.column("Time of Visit:", anchor= W, width = 120)
    tv1.column("Message:", anchor= W, width = 120)
                
    tv1.heading("#0", text="CUSTOMER NO.", anchor=W)
    tv1.heading("Reservation No.", text="Reservation No.", anchor= W)
    tv1.heading("Fullname", text="Fullname", anchor= W)
    tv1.heading("Email", text="Email", anchor= W)
    tv1.heading("Overnight", text="Overnight", anchor= W)
    tv1.heading("Cottage Size", text="Cottage Size", anchor= W)
    tv1.heading("Contact No.", text="Contact No.", anchor= W)
    tv1.heading("Time of Visit:", text="Time of Visit:", anchor= W)
    tv1.heading("Message:", text="Message:", anchor= W)


    for each_cell in range(2,(excel_activate.max_row)+1):
        tv1.insert(parent='', index="end", text=str(each_cell), values=(excel_activate["A"+str(each_cell)].value,excel_activate["B"+str(each_cell)].value,excel_activate["C"+str(each_cell)].value,excel_activate["D"+str(each_cell)].value,excel_activate["E"+str(each_cell)].value,excel_activate["F"+str(each_cell)].value,excel_activate["G"+str(each_cell)].value,excel_activate["H"+str(each_cell)].value))
    tv1.pack()
    view_frame.mainloop()



help= Label(root, text= "How May i Help You?" ,font =("Arial", 10, "bold"),bg ='lightblue', bd= 9)
help.place(x= 20, y=350,height=15)

#BUTTONS===============================================================================================================================================================================================================================


btn1 = Button(root, text = "RESERVE", font =("Arial", 13, "bold"),bg= "skyblue",fg = "black", bd= 9,command=lambda:reserve())
btn1.place(x= 20, y=380,height=35)

btn2 = Button(root, text = "VIEW RESERVATIONS", font =("Arial", 13, "bold"),bg= "skyblue",fg = "black",command=lambda:view2() , bd= 9)
btn2.place(x= 20, y=420,height=35)

root.mainloop()